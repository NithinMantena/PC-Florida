#!/usr/bin/env python3
"""
Florida P&C Market Explorer — ETL / ingestion.

Scans the data folder for FLOIR QUASR / Quarterly-MIR residential market-share
workbooks (both file types), normalizes them into a single longitudinal dataset,
and writes a self-contained `web/data.js` that the static frontend loads (no
server needed) plus a human-readable validation report.

Design notes (see Florida_PC_PRD.md):
  * Columns are mapped by HEADER TEXT, never by position. Column counts drift
    across quarters (Type B: 19 -> 29 -> 34 cols; Type A: 9 -> 39 -> 57 cols);
    later quarters merely append claims / lawsuit metrics. Matching on header
    text makes the parser immune to that drift and to the date wording changing
    inside each header.
  * The reporting quarter is taken from the filename token (e.g. `2022q2`),
    which is stable; the title's date range wording is not (it changes from
    "4/1/2022-6/30/2022" to "January 1, 2026 - March 31, 2026").
  * NAIC code is the stable entity key and is kept as a string.
  * Suppressed cells (a literal ".") become null + is_suppressed, never 0.
  * The footer "Total" row is detected and excluded from entity rollups, but is
    retained as a published checksum for validation.
  * Re-running is idempotent and, when the same period appears twice, the file
    with the newest pull-timestamp in its name wins.

Run:  python etl/ingest.py
"""

from __future__ import annotations

import glob
import json
import os
import re
import sys
from collections import defaultdict
from datetime import date

import openpyxl

# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #
ETL_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.dirname(ETL_DIR)                 # the "Florida P&C" folder
WEB_DIR = os.path.join(DATA_DIR, "web")
OUT_DATA_JS = os.path.join(WEB_DIR, "data.js")
OUT_REPORT = os.path.join(DATA_DIR, "validation_report.txt")

# --------------------------------------------------------------------------- #
# Header -> canonical metric mapping
#
# Each entry: (canonical_key, predicate(normalized_header) -> bool).
# Predicates are evaluated in order; the FIRST match wins, so more specific
# rules (e.g. "exclude wind", "due to hurricane") must precede general ones.
# `h` is the header lowercased with all whitespace collapsed to single spaces.
# --------------------------------------------------------------------------- #

def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s).replace("\n", " ")).strip().lower() if s is not None else ""


# ---- Type B (company x policy type) -------------------------------------- #
# Returns canonical metric key or None.
B_METRIC_RULES = [
    # --- policies in force (count). Anchor on "number of policies in force"
    #     at the START so we don't swallow the exposure/premium headers, which
    #     also contain the phrase "...for policies in force that exclude wind". ---
    ("pif_excl_wind", lambda h: h.startswith("number of policies in force") and "exclude wind" in h),
    ("pif_incl_wind", lambda h: h.startswith("number of policies in force") and "include wind" in h),
    ("pif", lambda h: h.startswith("total number of policies in force")),
    # --- policy flow (count) ---
    ("cancelled_hurricane", lambda h: "canceled due to hurricane" in h),
    ("cancelled", lambda h: "policies canceled" in h),
    ("nonrenewed_hurricane", lambda h: "nonrenewed due to hurricane" in h),
    ("nonrenewed", lambda h: "policies nonrenewed" in h),
    ("transferred_out", lambda h: "transferred to other insurers" in h),
    ("new_written", lambda h: "new policies written" in h),
    ("received_in", lambda h: "received from other insurers" in h),
    # --- exposure / TIV (dollars) ---
    ("tiv_excl_wind", lambda h: "value of exposure" in h and "exclude wind" in h),
    ("tiv_incl_wind", lambda h: "value of exposure" in h and "include wind" in h),
    ("tiv", lambda h: "value of exposure" in h),
    # --- direct premium written (dollars) ---
    ("dpw_excl_wind", lambda h: "direct premium written" in h and "exclude wind" in h),
    ("dpw_incl_wind", lambda h: "direct premium written" in h and "include wind" in h),
    ("dpw", lambda h: "direct premium written" in h),
    # --- claims (count; newer quarters only) ---
    ("claims_opened", lambda h: h.startswith("total number of claims opened")),
    ("claims_closed", lambda h: h.startswith("total number of claims closed")),
    ("claims_pending", lambda h: h.startswith("total number of claims pending")),
    # --- lawsuits (count; newest quarters only) ---
    ("lawsuits_closed_consumer", lambda h: "lawsuits closed with consideration" in h),
    ("lawsuits_closed", lambda h: h.startswith("number of lawsuits closed")),
    ("lawsuits_opened", lambda h: h.startswith("number of lawsuits opened")),
    ("lawsuits_open_end", lambda h: "lawsuits open at end" in h),
]

# Anchor / dimension columns for Type B
B_DIM_RULES = [
    ("naic", lambda h: h == "naic code"),
    ("company", lambda h: h == "company name"),
    ("policy_type", lambda h: h == "policy type"),
]

# ---- Type A (company x commercial/personal) ------------------------------ #
A_METRIC_RULES = [
    ("a_pif_commercial", lambda h: "policies in force that are commercial" in h),
    ("a_pif_personal", lambda h: "policies in force that are personal" in h),
    ("a_pif", lambda h: h.startswith("total number of policies in force")),
    ("a_dpw_commercial", lambda h: "direct premium written for policies in force that are commercial" in h),
    ("a_dpw_personal", lambda h: "direct premium written for policies in force that are personal" in h),
    ("a_dpw", lambda h: h.startswith("total direct premium written for policies in force")),
]
A_DIM_RULES = [
    ("rank", lambda h: h.startswith("rank by total")),
    ("company", lambda h: h == "company name"),
    ("naic", lambda h: h == "naic code"),
]


def classify(header_norm: str, rules) -> str | None:
    for key, pred in rules:
        if pred(header_norm):
            return key
    return None


# --------------------------------------------------------------------------- #
# Value cleaning
# --------------------------------------------------------------------------- #

def clean_number(v):
    """Return (value_or_None, is_suppressed). Handles '.', '$1,234', floats."""
    if v is None:
        return None, False
    if isinstance(v, (int, float)):
        return float(v), False
    s = str(v).strip()
    if s == "" :
        return None, False
    if s == ".":
        return None, True            # suppressed / redacted, NOT zero
    s = re.sub(r"[\$,\s]", "", s)
    if s in ("", "-"):
        return None, False
    try:
        return float(s), False
    except ValueError:
        return None, False


# --------------------------------------------------------------------------- #
# Period / quarter helpers
# --------------------------------------------------------------------------- #
QUARTER_END = {1: (3, 31), 2: (6, 30), 3: (9, 30), 4: (12, 31)}


def parse_period_from_filename(fname: str):
    m = re.search(r"(\d{4})q([1-4])", fname.lower())
    if not m:
        return None
    yyyy, q = int(m.group(1)), int(m.group(2))
    mm, dd = QUARTER_END[q]
    return f"{yyyy}Q{q}", date(yyyy, mm, dd).isoformat()


def parse_timestamp_token(fname: str) -> str:
    """The trailing pull-timestamp token, e.g. 20221220t133802, for dedupe."""
    m = re.search(r"(\d{8}t\d{6})", fname.lower())
    return m.group(1) if m else ""


# --------------------------------------------------------------------------- #
# Workbook reading
# --------------------------------------------------------------------------- #

def load_sheet(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    sheet_name = wb.sheetnames[0]
    wb.close()
    return sheet_name, rows


def find_header_row(rows, anchor="company name", max_scan=10):
    """Locate the header row by scanning for the 'Company name' anchor."""
    for i, r in enumerate(rows[:max_scan]):
        if any(_norm(c) == anchor for c in r):
            return i
    return None


def title_and_pulled(rows):
    title = _norm(rows[0][0]) if rows and rows[0] else ""
    pulled = ""
    if len(rows) > 1 and rows[1] and rows[1][0]:
        m = re.search(r"data pulled on (.+?)\)", str(rows[1][0]), re.I)
        pulled = m.group(1).strip() if m else str(rows[1][0]).strip()
    return rows[0][0] if (rows and rows[0]) else "", pulled


def is_total_row(name_cell) -> bool:
    return _norm(name_cell) == "total"


# --------------------------------------------------------------------------- #
# Policy-type parsing (line + wind-only) — discovered, not hardcoded
# --------------------------------------------------------------------------- #

def parse_policy_type(pt: str):
    p = pt.strip()
    low = p.lower()
    if low.startswith("commercial residential"):
        line = "Commercial Residential"
    elif low.startswith("personal residential"):
        line = "Personal Residential"
    else:
        line = "Other"            # surfaced as a data-quality warning
    is_wind_only = "wind only" in low
    return line, is_wind_only, (line == "Other")


# --------------------------------------------------------------------------- #
# Parse one Type B workbook
# --------------------------------------------------------------------------- #

def parse_type_b(path, period, period_end, warnings):
    sheet, rows = load_sheet(path)
    hdr_i = find_header_row(rows)
    if hdr_i is None:
        warnings.append(f"[B] {os.path.basename(path)}: no header row found; skipped")
        return [], None, None
    header = rows[hdr_i]
    # map column index -> canonical key (dims + metrics)
    colmap = {}
    for ci, cell in enumerate(header):
        hn = _norm(cell)
        if not hn:
            continue
        key = classify(hn, B_DIM_RULES) or classify(hn, B_METRIC_RULES)
        if key:
            colmap[ci] = key
    metric_cols = {ci: k for ci, k in colmap.items()
                   if k not in ("naic", "company", "policy_type")}
    # locate dim columns
    inv = {k: ci for ci, k in colmap.items()}
    if not all(k in inv for k in ("naic", "company", "policy_type")):
        warnings.append(f"[B] {os.path.basename(path)}: missing dim columns; skipped")
        return [], None, None

    facts = []
    total_row = None
    for r in rows[hdr_i + 1:]:
        if r is None or all(c is None for c in r):
            continue
        name = r[inv["company"]]
        rec = {}
        for ci, mkey in metric_cols.items():
            val, sup = clean_number(r[ci] if ci < len(r) else None)
            rec[mkey] = val
            if sup:
                rec.setdefault("_suppressed", []).append(mkey)
        if is_total_row(name):
            total_row = rec
            continue
        naic = ("" if r[inv["naic"]] is None else str(r[inv["naic"]]).strip())
        pt = ("" if r[inv["policy_type"]] is None else str(r[inv["policy_type"]]).strip())
        if not naic and not pt:
            continue
        line, wind_only, unknown = parse_policy_type(pt)
        if unknown:
            warnings.append(f"[B] {period}: unrecognized policy-type prefix: {pt!r}")
        fact = {
            "p": period,
            "naic": naic,
            "company": ("" if name is None else str(name).strip()),
            "pt": pt,
            "line": line,
            "wind_only": wind_only,
        }
        fact.update({k: v for k, v in rec.items() if k != "_suppressed"})
        if "_suppressed" in rec:
            fact["_sup"] = rec["_suppressed"]
        facts.append(fact)
    return facts, total_row, sorted(set(metric_cols.values()))


# --------------------------------------------------------------------------- #
# Parse one Type A workbook
# --------------------------------------------------------------------------- #

def parse_type_a(path, period, period_end, warnings):
    sheet, rows = load_sheet(path)
    hdr_i = find_header_row(rows)
    if hdr_i is None:
        warnings.append(f"[A] {os.path.basename(path)}: no header row found; skipped")
        return [], None
    header = rows[hdr_i]
    colmap = {}
    for ci, cell in enumerate(header):
        hn = _norm(cell)
        if not hn:
            continue
        key = classify(hn, A_DIM_RULES) or classify(hn, A_METRIC_RULES)
        if key:
            colmap[ci] = key
    inv = {k: ci for ci, k in colmap.items()}
    if not all(k in inv for k in ("naic", "company")):
        warnings.append(f"[A] {os.path.basename(path)}: missing dim columns; skipped")
        return [], None
    metric_cols = {ci: k for ci, k in colmap.items()
                   if k not in ("naic", "company", "rank")}

    facts = []
    total_row = None
    for r in rows[hdr_i + 1:]:
        if r is None or all(c is None for c in r):
            continue
        name = r[inv["company"]]
        rec = {}
        for ci, mkey in metric_cols.items():
            val, _ = clean_number(r[ci] if ci < len(r) else None)
            rec[mkey] = val
        if is_total_row(name):
            total_row = rec
            continue
        naic = ("" if r[inv["naic"]] is None else str(r[inv["naic"]]).strip())
        if not naic:
            continue
        fact = {"p": period, "naic": naic,
                "company": ("" if name is None else str(name).strip())}
        fact.update(rec)
        facts.append(fact)
    return facts, total_row


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def discover_files():
    """Return {('B'|'A', period): path} keeping newest pull-timestamp per slot."""
    chosen = {}
    for path in glob.glob(os.path.join(DATA_DIR, "*.xlsx")):
        base = os.path.basename(path)
        low = base.lower()
        if "by_company_and_policy_type" in low:
            ftype = "B"
        elif "by_company_and_commercial_personal" in low:
            ftype = "A"
        else:
            continue
        per = parse_period_from_filename(base)
        if not per:
            continue
        period, _ = per
        key = (ftype, period)
        ts = parse_timestamp_token(base)
        if key not in chosen or ts > chosen[key][1]:
            chosen[key] = (path, ts)
    return {k: v[0] for k, v in chosen.items()}


def main():
    files = discover_files()
    if not files:
        print("No FLOIR workbooks found in", DATA_DIR)
        sys.exit(1)

    warnings = []
    facts_b, facts_a = [], []
    periods = {}                         # period -> metadata
    checksums = []                       # validation rows
    metrics_seen = set()

    for (ftype, period), path in sorted(files.items(), key=lambda kv: (kv[0][1], kv[0][0])):
        period_key, period_end = parse_period_from_filename(os.path.basename(path))
        sheet, rows = load_sheet(path)
        title_raw, pulled = title_and_pulled(rows)
        pmeta = periods.setdefault(period_key, {
            "period": period_key, "period_end": period_end,
            "pulled_at": pulled, "has_A": False, "has_B": False,
            "title": title_raw,
        })
        pmeta["pulled_at"] = pmeta["pulled_at"] or pulled

        if ftype == "B":
            f, total, mets = parse_type_b(path, period_key, period_end, warnings)
            facts_b.extend(f)
            pmeta["has_B"] = True
            if mets:
                metrics_seen.update(mets)
            # checksum: sum body PIF vs Total-row PIF
            body_pif = sum(x.get("pif") or 0 for x in f)
            tot_pif = (total or {}).get("pif")
            checksums.append(("B", period_key, "pif", body_pif, tot_pif,
                              len(f), os.path.basename(path)))
        else:
            f, total = parse_type_a(path, period_key, period_end, warnings)
            facts_a.extend(f)
            pmeta["has_A"] = True
            body_pif = sum(x.get("a_pif") or 0 for x in f)
            tot_pif = (total or {}).get("a_pif")
            checksums.append(("A", period_key, "a_pif", body_pif, tot_pif,
                              len(f), os.path.basename(path)))

    # ---- build dimension tables ----------------------------------------- #
    companies = {}
    for fct in facts_b + facts_a:
        naic = fct["naic"]
        c = companies.setdefault(naic, {"naic": naic, "names": {}, "periods": set()})
        nm = fct.get("company", "")
        if nm:
            c["names"][nm] = c["names"].get(nm, 0) + 1
        c["periods"].add(fct["p"])

    period_order = sorted(periods.keys(), key=lambda p: (int(p[:4]), int(p[-1])))
    companies_out = {}
    for naic, c in companies.items():
        # most frequent name = display label
        display = max(c["names"].items(), key=lambda kv: kv[1])[0] if c["names"] else naic
        pers = sorted(c["periods"], key=lambda p: (int(p[:4]), int(p[-1])))
        companies_out[naic] = {
            "naic": naic,
            "name": display,
            "names": sorted(c["names"].keys()),
            "first": pers[0] if pers else None,
            "last": pers[-1] if pers else None,
        }

    policy_types = {}
    for fct in facts_b:
        pt = fct["pt"]
        if pt and pt not in policy_types:
            line, wind_only, _ = parse_policy_type(pt)
            policy_types[pt] = {"policy_type": pt, "line": line, "is_wind_only": wind_only}

    data = {
        "generated_at": _now(),
        "periods": [periods[p] for p in period_order],
        "companies": companies_out,
        "policy_types": policy_types,
        "metrics_b": sorted(metrics_seen),
        "facts_b": facts_b,
        "facts_a": facts_a,
    }

    os.makedirs(WEB_DIR, exist_ok=True)
    with open(OUT_DATA_JS, "w", encoding="utf-8") as fh:
        fh.write("// AUTO-GENERATED by etl/ingest.py — do not edit by hand.\n")
        fh.write("window.FL_DATA = ")
        json.dump(data, fh, ensure_ascii=False, separators=(",", ":"))
        fh.write(";\n")

    write_report(files, periods, period_order, checksums, warnings,
                 facts_b, facts_a, companies_out, policy_types, metrics_seen)

    print(f"OK  periods={len(period_order)}  facts_B={len(facts_b)}  "
          f"facts_A={len(facts_a)}  companies={len(companies_out)}  "
          f"policy_types={len(policy_types)}")
    print(f"    wrote {OUT_DATA_JS}")
    print(f"    wrote {OUT_REPORT}")
    if warnings:
        print(f"    {len(warnings)} warning(s) — see report")


def _now():
    from datetime import datetime
    return datetime.now().isoformat(timespec="seconds")


def write_report(files, periods, period_order, checksums, warnings,
                 facts_b, facts_a, companies, policy_types, metrics_seen):
    L = []
    L.append("FLORIDA P&C — INGEST VALIDATION REPORT")
    L.append(f"generated: {_now()}")
    L.append(f"input dir: {DATA_DIR}")
    L.append("")
    L.append(f"files ingested: {len(files)}")
    L.append(f"periods: {len(period_order)}  ({period_order[0]} .. {period_order[-1]})")
    L.append(f"Type-B facts (company x policy type): {len(facts_b)}")
    L.append(f"Type-A facts (company x comm/personal): {len(facts_a)}")
    L.append(f"distinct companies (NAIC): {len(companies)}")
    L.append(f"distinct policy types: {len(policy_types)}")
    L.append(f"metrics captured (Type B): {', '.join(sorted(metrics_seen))}")
    L.append("")
    L.append("PERIOD COVERAGE (A=comm/personal, B=policy type):")
    for p in period_order:
        m = periods[p]
        L.append(f"  {p}  A={'Y' if m['has_A'] else '-'}  B={'Y' if m['has_B'] else '-'}"
                 f"  pulled={m['pulled_at']}")
    L.append("")
    L.append("TOTAL-ROW CHECKSUM (body sum vs published Total row, PIF):")
    L.append(f"  {'type':4} {'period':7} {'metric':6} {'body_sum':>14} {'total_row':>14} {'diff':>10} {'rows':>5}")
    for ftype, period, metric, body, tot, n, fn in sorted(checksums, key=lambda x: (x[1], x[0])):
        diff = "" if tot is None else f"{body - tot:,.0f}"
        tots = "" if tot is None else f"{tot:,.0f}"
        flag = ""
        if tot is not None and tot != 0 and abs(body - tot) / tot > 0.0001:
            flag = "  <-- MISMATCH"
        L.append(f"  {ftype:4} {period:7} {metric:6} {body:>14,.0f} {tots:>14} {diff:>10} {n:>5}{flag}")
    L.append("")
    if warnings:
        L.append(f"WARNINGS ({len(warnings)}):")
        for w in warnings:
            L.append(f"  - {w}")
    else:
        L.append("WARNINGS: none")
    L.append("")
    L.append("POLICY TYPES DISCOVERED:")
    for pt in sorted(policy_types):
        d = policy_types[pt]
        L.append(f"  [{d['line'][:4]}{' WIND' if d['is_wind_only'] else '    '}] {pt}")
    with open(OUT_REPORT, "w", encoding="utf-8") as fh:
        fh.write("\n".join(L) + "\n")


if __name__ == "__main__":
    main()
